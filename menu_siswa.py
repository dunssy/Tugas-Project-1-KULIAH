import openpyxl
import os

FILE_NAME = 'data_beasiswa.xlsx'

def create_sheet_if_not_exists (workbook, sheet_name, header=None):
    """Buat sheet baru jika tidak ada dan tambahkan header jika
diperlukan."""

    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet (sheet_name)
        if header:
            sheet.append (header)
    return workbook [sheet_name]

# Fungsi untuk menambah siswa
def tambah_siswa ():
    nisn_siswa = input ("Masukkan nomor siswa: ")
    nama_siswa = input ("Masukkan nama siswa: ")
    no_hp = input ("Masukkan nomor HP siswa: ")
    alamat = input ("Masukkan alamat siswa: ")

    if not os.path.exists (FILE_NAME):
        workbook = openpyxl.Workbook ()
    else:
        workbook = openpyxl.load_workbook (FILE_NAME)
    # Cek atau buat sheet siswa
    sheet = create_sheet_if_not_exists (workbook, 'Siswa', ['No siswa', 'Nama siswa', 'No HP', 'Alamat']) 
    sheet.append([nisn_siswa, nama_siswa, no_hp, alamat])
    workbook.save (FILE_NAME)
    print("siswa berhasil ditambahkan.")

# Fungsi untuk menampilkan siswa
def tampil_siswa ():
    if not os.path.exists (FILE_NAME):
        print ("File data perpustakaan tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook (FILE_NAME)
    if 'Siswa' not in workbook.sheetnames:
        print("Sheet siswa belum ada.")
        return

    sheet = workbook['Siswa']
    if sheet.max_row == 1:
        print ("Belum ada data siswa.")
        return

    print("\nDaftar siswa:")
    for row in sheet.iter_rows (min_row=2, values_only=True):
        print (row)
    workbook.close ()

#Fungsi untuk mengedit siswa
def edit_siswa ():
    nisn_siswa = input ("Masukkan nomor siswa yang ingin diedit: ")

    if not os.path.exists (FILE_NAME):
        print ("File data perpustakaan tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if 'siswa' not in workbook.sheetnames:
        print("Sheet siswa belum ada.")
        return

    sheet = workbook['Siswa']
    for row in sheet.iter_rows (min_row=2):
        if row [0].value == nisn_siswa:
            nama_siswa = input("Masukkan nama siswa baru: ")
            no_hp = input("Masukkan nomor HP siswa baru: ")
            alamat = input("Masukkan alamat siswa baru: ")
            row[1].value = nama_siswa
            row[2].value = no_hp
            row[3].value = alamat
            print ("siswa berhasil diedit.")
            workbook.save (FILE_NAME)
            return

    print ("siswa tidak ditemukan.")
    workbook.save (FILE_NAME)

# Fungsi untuk menghapus siswa
def hapus_siswa ():
    nisn_siswa = input ("Masukkan nomor siswa yang ingin dihapus: ")

    if not os.path.exists (FILE_NAME):
        print ("File data perpustakaan tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if 'siswa' not in workbook.sheetnames:
        print("Sheet siswa belum ada.")
        return
    
    sheet = workbook['Siswa']
    rows_to_delete = []
    for row_index, row in enumerate(sheet.iter_rows (min_row=2), start=2):
        if row [0].value == nisn_siswa:
            rows_to_delete.append (row_index)

    if rows_to_delete:  
        for row_index in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows (row_index)
        print ("siswa berhasil dihapus.")
    else:
        print ("siswa tidak ditemukan.")

    workbook.save (FILE_NAME)

def menu_siswa ():
    while True:
        print ("\nMenu siswa:")
        print ("1. Tambah siswa")
        print ("2. Tampil siswa")
        print ("3. Edit siswa")
        print ("4. Hapus siswa")
        print ("5. Kembali ke Menu Utama")
        sub_pilihan = input("Pilih menu: ")
        if sub_pilihan == '1':
            tambah_siswa()
        elif sub_pilihan == '2':
            tampil_siswa()
        elif sub_pilihan =='3':
            edit_siswa()
        elif sub_pilihan == '4':
            hapus_siswa()
        elif sub_pilihan == '5':
            break

def daftar_beasiswa():
    nisn = input("Masukkan NISN siswa: ")
    nama_beasiswa = input("Masukkan nama beasiswa yang didaftar: ")

    # Cek file
    if not os.path.exists(FILE_NAME):
        print("File data tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)

    # Cek sheet beasiswa
    if 'beasiswa' not in workbook.sheetnames:
        print("Belum ada data beasiswa.")
        return

    sheet_beasiswa = workbook['beasiswa']

    # Cari kuota beasiswa
    kuota = None
    for row in sheet_beasiswa.iter_rows(min_row=2, values_only=True):
        if row[0] == nama_beasiswa:
            kuota = int(row[1])
            break

    if kuota is None:
        print("Beasiswa tidak ditemukan.")
        return

    # Buat sheet pendaftaran jika belum ada
    sheet_daftar = create_sheet_if_not_exists(
        workbook,
        'pendaftaran',
        ['NISN', 'Nama Beasiswa', 'Status']
    )

    # Hitung jumlah penerima
    diterima = 0
    for row in sheet_daftar.iter_rows(min_row=2, values_only=True):
        if row[1] == nama_beasiswa and row[2] == "DITERIMA":
            diterima += 1

    # Tentukan status
    if diterima < kuota:
        status = "DITERIMA"
    else:
        status = "DITOLAK"

    # Simpan pendaftar
    sheet_daftar.append([nisn, nama_beasiswa, status])
    workbook.save(FILE_NAME)

    print(f"Pendaftaran disimpan. Status: {status}")
