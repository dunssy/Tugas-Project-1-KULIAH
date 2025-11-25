import openpyxl
import os
from datetime import datetime

FILE_NAME = 'data_beasiswa.xlsx' 

def create_sheet_if_not_exists (workbook, sheet_name, header=None):
    """Buat sheet baru jika tidak ada dan tambahkan header jika
diperlukan."""
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
        if header:
            sheet.append(header)
    return workbook [sheet_name]

def memisahkan_string (kode): 
    if len(kode) != 6:
        print ("Input harus terdiri dari 6 karakter.")
        exit ()
    jenis_beasiswa = kode [0:3]
    nomor_beasiswa = kode [3:6]
    return jenis_beasiswa, nomor_beasiswa

def tentukan_jenis_beasiswa(jenis_beasiswa):
    if jenis_beasiswa ==  "B01":
        return "Beasiswa Negara"
    elif jenis_beasiswa == "B02":
        return "Beasiswa Swasta"
    elif jenis_beasiswa == "B03":
        return "Beasiswa Perusahaan"
    else:
        return "Jenis beasiswa tidak dikenal"
    
# Fungsi untuk menambah beasiswa
def tambah_beasiswa ():
    kode = input ("Masukkan kode beasiswa (3 karakter jenis_beasiswa, 3 karakter nomor beasiswa, contoh: B01001): ")
    nama = input("Masukkan nama beasiswa: ")
    pemberi = input ("Masukkan pemberi beasiswa: ")
    kouta = input ("Masukan Kouta Beasiswa: ")
    jenis_beasiswa_kode, nomor_beasiswa = memisahkan_string (kode)
    jenis_beasiswa_nama = tentukan_jenis_beasiswa(jenis_beasiswa_kode)

    if not os.path.exists (FILE_NAME):
        workbook = openpyxl.Workbook ()
    else:
        workbook = openpyxl.load_workbook (FILE_NAME)

    sheet = create_sheet_if_not_exists (workbook, 'Beasiswa', ['Kode Beasiswa','Nama Beasiswa', 'Pemberi beasiswa', 'Jenis Beasiswa', 'Nomor Beasiswa', 'Kouta Beasiswa', 'Status'])
    sheet.append ([kode, nama, pemberi, jenis_beasiswa_nama, nomor_beasiswa, kouta,'Tersedia'])
    workbook.save (FILE_NAME)
    print("Beasiswa berhasil ditambahkan.")

# Fungsi untuk menampilkan beasiswa
def tampil_beasiswa():
    if not os.path.exists (FILE_NAME):
        print("File data beasiswa tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook (FILE_NAME)
    if 'Beasiswa' not in workbook.sheetnames:
        print("Sheet Beasiswa belum ada.")
        return

    sheet = workbook ['Beasiswa']
    if sheet.max_row == 1:
        print ("Belum ada data beasiswa.")
        return

    print("\nDaftar Beasiswa:")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print (row)
    workbook.close ()

# Fungsi untuk mengedit beasiswa
def edit_beasiswa():
    kode = input ("Masukkan kode beasiswa yang ingin diedit: ")

    if not os.path.exists (FILE_NAME): 
        print("File data beasiswa tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook (FILE_NAME)
    if 'Beasiswa' not in workbook.sheetnames:
        print("Sheet Beasiswa belum ada.")
        return

    sheet = workbook['Beasiswa']
    for row in sheet.iter_rows (min_row=2):
        if row[0].value == kode:
            nama = input ("Masukkan nama beasiswa baru: ")
            pemberi = input ("Masukkan pemberi beasiswa baru: ")
            row[1].value = nama
            row [2].value = pemberi
            print("Beasiswa berhasil diedit.")
            workbook.save (FILE_NAME)
            return

    print("Beasiswa tidak ditemukan.")
    workbook.save (FILE_NAME)

# Fungsi untuk menghapus beasiswa
def hapus_beasiswa():
    kode = input ("Masukkan kode beasiswa yang ingin dihapus: ")

    if not os.path.exists (FILE_NAME):
        print ("File data perpustakaan tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook (FILE_NAME)
    if 'Beasiswa' not in workbook.sheetnames:
        print("Sheet Beasiswa belum ada.")
        return

    sheet = workbook['Beasiswa']
    rows_to_delete = []
    for row_index, row in enumerate(sheet.iter_rows (min_row=2), start=2):
        if row[0].value == kode:
            rows_to_delete.append(row_index)

    if rows_to_delete:
        for row_index in sorted(rows_to_delete, reverse=True) :
            sheet.delete_rows (row_index)
        print ("Beasiswa berhasil dihapus.")
    else:
        print("Beasiswa tidak ditemukan.")

    workbook.save (FILE_NAME)

def menu_beasiswa () :
    while True :
        print ("\nMenu Beasiswa:")
        print ("1. Tambah Beasiswa")
        print ("2. Tampil Beasiswa")
        print ("3. Edit Beasiswa")
        print ("4. Hapus Beasiswa")
        print ("5. Kembali ke Menu Utama")
        sub_pilihan = input ("Pilih menu: ")
        
        if sub_pilihan == '1':
            tambah_beasiswa()
        elif sub_pilihan == '2':
            tampil_beasiswa()
        elif sub_pilihan == '3':
            edit_beasiswa()
        elif sub_pilihan == '4':
            hapus_beasiswa()
        elif sub_pilihan == '5':
            break