import openpyxl
import os
from datetime import datetime

FILE_NAME = 'data_beasiswa.xlsx'

def create_sheet_if_not_exists (workbook, sheet_name, header=None):
    """Buat sheet baru jika tidak ada dan tambahkan header jika diperlukan."""
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
        if header:
            sheet.append(header)
    return workbook [sheet_name]

def laporan_beasiswa():
    nisn_siswa = input ("Masukkan nisn siswa: ")
    kode_beasiswa = input ("Masukkan kode beasiswa: ")
    laporan_beasiswa = datetime.today().strftime('%Y-%m-%d')

    if not os.path.exists (FILE_NAME):
        print ("Data beasiswa belum ada.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if 'Siswa' not in workbook.sheetnames or 'Beasiswa' not in workbook.sheetnames or 'Pemberian' not in workbook.sheetnames:
        print ("Data beasiswa, siswa, atau pemberian tidak ada.")
        return

    pemberian_sheet = workbook ['Pemberian']
    buku_sheet = workbook ['Beasiswa']
    
    pemberian_valid = False
    for row in pemberian_sheet.iter_rows (min_row=2):
        if row[0].value == nisn_siswa and row[1].value == kode_beasiswa:
            pemberian_valid = True
            pemberian_sheet.delete_rows(row[0].row)
            break

    if not pemberian_valid:
        print ("Pemberian beasiswa tidak ditemukan.")
        return
   
    laporan_sheet = create_sheet_if_not_exists (workbook, 'Laporan', ['NISN Siswa', 'Kode Beasiswa', 'Tanggal Laporan'])
    laporan_sheet.append([nisn_siswa, kode_beasiswa, laporan_beasiswa])

    for row in buku_sheet.iter_rows (min_row=2):
        if row[0].value == kode_beasiswa:
            row[6].value = "Tersedia"
    
    workbook.save (FILE_NAME)
    print("Beasiswa berhasil dikembalikan.")

def tampil_data_laporan():
    if not os.path.exists (FILE_NAME):
        print ("File data beasiswa tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook (FILE_NAME)

    if 'Laporan' not in workbook.sheetnames:
        print("Sheet Laporan belum ada.")
        return
    
    sheet_laporan = workbook ['Laporan']
    if sheet_laporan.max_row == 1:
        print("Belum ada data pengembalian.")
        return

    print("\nDaftar Laporan:")
    for row in sheet_laporan.iter_rows (min_row=2, values_only=True):
        print (row)
    workbook.close ()

def menu_laporan() :
    while True:
        print("\nMenu Laporan:")
        print("1. Laporan Beasiswa")
        print("2. Tampil data pengembalian")
        print("3. Kembali ke Menu Utama")
        sub_pilihan = input ("Pilih menu: ")
        if sub_pilihan == '1':
            laporan_beasiswa()
        elif sub_pilihan == '2':
            tampil_data_laporan()
        elif sub_pilihan == '3':
            break