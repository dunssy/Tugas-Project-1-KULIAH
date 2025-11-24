import openpyxl
import os
from datetime import datetime

FILE_NAME = 'data_beasiswa.xlsx'

def create_sheet_if_not_exists (workbook, sheet_name, header=None):
    """Buat sheet baru jika tidak ada dan tambahkan header jika diperlukan."""
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet (sheet_name)
        if header:
            sheet.append (header)
    return workbook [sheet_name]

def pemberian_beasiswa():

    nisn_siswa = input ("Masukkan nisn siswa: ")
    kode_beasiswa = input ("Masukkan kode beasiswa: ")
    tanggal_terima = datetime.today().strftime('%Y-%m-%d')
    
    if not os.path.exists (FILE_NAME):
        print ("Data beasiswa belum ada.")
        return

    workbook = openpyxl.load_workbook (FILE_NAME)
    if 'Siswa' not in workbook.sheetnames or 'Beasiswa' not in workbook.sheetnames:
        print ("Data beasiswa atau siswa tidak ada.")
        return

    siswa_sheet = workbook ['Siswa']
    beasiswa_sheet = workbook ['Beasiswa']
    
    siswa_valid = False
    for row in siswa_sheet.iter_rows (min_row=2):
        if row [0].value == nisn_siswa:
            siswa_valid = True
            break

    beasiswa_valid = False
    for row in beasiswa_sheet.iter_rows (min_row=2):
        if row[0].value == kode_beasiswa and row[6].value == "Tersedia":
            beasiswa_valid = True
            break

    if not siswa_valid:
        print("Siswa tidak terdaftar.")
        return

    if not beasiswa_valid:
        print("Beasiswa tidak tersedia untuk diterima.")
        return

    pemberian_sheet = create_sheet_if_not_exists (workbook, 'Pemberian', ['NISN Siswa', 'Kode Beasiswa', 'Tanggal Terima'])
    pemberian_sheet.append([nisn_siswa, kode_beasiswa, tanggal_terima])

    for row in beasiswa_sheet.iter_rows (min_row=2):
        if row [0].value == kode_beasiswa:
            row[6].value = "Diterima"

    workbook.save (FILE_NAME)
    print("Beasiswa berhasil diterima.")

def tampil_data_pemberian ():
    if not os.path.exists (FILE_NAME):
        print("File data beasiswa tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook (FILE_NAME)

    if 'Pemberian' not in workbook.sheetnames:
        print("Sheet Pemberian belum ada.")
        return

    sheet_penerimaan = workbook ['Pemberian']
    if sheet_penerimaan.max_row == 1:
        print("Belum ada data pemberian.")
        return

    print("\nDaftar Pemberian:")
    for row in sheet_penerimaan.iter_rows (min_row=2, values_only=True) :
        print (row)
    workbook.close ()

def menu_pemberian():
    while True:
        print("\nMenu Pemberian:")
        print("1. Terima Beasiswa")
        print("2. Tampil data pemberian")
        print("3. Kembali ke Menu Utama")
        sub_pilihan = input("Pilih menu: ")
        if sub_pilihan == '1':
            pemberian_beasiswa()
        elif sub_pilihan == '2':
            tampil_data_pemberian()
        elif sub_pilihan == '3':
            break